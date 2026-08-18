"""
Foodstuffs route generation
============================

Phase 1 of a two-phase vehicle routing approach: generate a large, diverse
pool of capacity-feasible candidate delivery routes (Warehouse -> stores ->
Warehouse), each tagged with its driving time, unload time, total time and
dollar cost. A later set-covering/partitioning optimisation (not part of
this module) picks the best subset of these candidates to actually run.

Business rules encoded here
----------------------------
- Truck capacity: 16 pallets. This is the ONLY hard constraint on which
  stores can be combined into one route -- it's a genuine physical limit
  and can never be exceeded.
- Unload time: 18 minutes per pallet, at the store.
- Driving time: taken directly from the provided duration matrix (already
  reflects average Auckland traffic conditions), in seconds -> minutes.
- Cost model: $220/hour for the first 4 hours of a route (driving +
  unloading combined), then $310 per hour or part thereof beyond 4 hours.
  The often-quoted "3.5 hours" is the planning target/average shift length,
  NOT a hard cutoff -- routes are allowed to run long, they just cost more.
  A handful of the busiest stores (Pak 'n Save, ~14 pallets) take over 4
  hours to unload alone; those routes are still generated, just flagged as
  expensive/overtime, per Foodstuffs' guidance to let the later
  optimisation weigh the trade-off rather than excluding them outright.
- Each store receives exactly one delivery per day -- so weekday and
  Saturday route pools are built as two entirely separate problems, from
  their respective pallet-demand figures.
- Stores with zero demand on a given day (e.g. closed on Saturday) need no
  delivery at all that day. They are excluded from route construction
  entirely -- no singleton route is forced for them, and they never appear
  in multi-store candidates -- and are reported separately rather than
  being treated as "never covered" by the pool.
- A generous sanity cutoff (default 6 hours) bounds candidate route
  duration purely to keep the pool computationally tractable -- it is not
  a business rule, just a practical limit on how long a "reasonable"
  candidate route can be. Singleton (one-store) routes are always kept
  regardless of this cutoff, since every store with nonzero demand must
  be covered by at least one candidate route no matter how expensive.

Pool construction
------------------
With 55 stores, exhaustively enumerating route combinations is impossible
(2^55 subsets). Instead this uses repeated randomized-greedy construction:
starting from many different random stores, greedily add the nearest
still-feasible next store (with a little randomness in which "nearby"
candidate gets picked, for diversity across restarts), snapshotting every
intermediate partial route along the way as its own candidate. This
produces a large, varied pool of routes of many different sizes touching
many different stores, which is what a downstream set-covering ILP needs
to have real options to choose between.

Every candidate is run through 2-opt (see `two_opt`) and *then* compared
against other candidates covering the same store set -- optimizing before
comparing, not after picking a "winner" -- since 2-opt is a local search
whose result depends on the starting order, so comparing raw (pre-2-opt)
construction times can discard an ordering that would have optimized
better than the one that "looked" best going in.
"""

import pandas as pd
import numpy as np
import random


CAPACITY_PALLETS = 16
UNLOAD_MIN_PER_PALLET = 18
NORMAL_RATE_HOURS = 4.0
NORMAL_RATE_PER_HOUR = 220.0
OVERTIME_RATE_PER_HOUR = 310.0
WAREHOUSE = "Warehouse"


def load_demand(xlsx_path, sheet_name="Required Supply"):
    """
    Load per-store pallet demand from the required_supply_weekday_saturday
    workbook.

    Returns
    -------
    (dict, dict)
        weekday_demand, saturday_demand -- each {store_name: pallets}.
    """
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name)
    weekday_demand = dict(zip(df["Supermarket"], df["Weekday"]))
    saturday_demand = dict(zip(df["Supermarket"], df["Saturday"]))
    return weekday_demand, saturday_demand


def load_duration_matrix(csv_path):
    """Load the Warehouse<->store driving-time matrix (seconds) as a DataFrame."""
    return pd.read_csv(csv_path, index_col=0)


def route_driving_minutes(route, dur):
    """Driving time (minutes) for Warehouse -> route[0] -> ... -> route[-1] -> Warehouse."""
    if not route:
        return 0.0
    t = dur.loc[WAREHOUSE, route[0]]
    for a, b in zip(route, route[1:]):
        t += dur.loc[a, b]
    t += dur.loc[route[-1], WAREHOUSE]
    return t / 60.0


def route_total_minutes(route, demand, dur):
    """Driving + unload time (minutes) for a route."""
    pallets = sum(demand[s] for s in route)
    return route_driving_minutes(route, dur) + UNLOAD_MIN_PER_PALLET * pallets


def route_cost(total_minutes):
    """
    $220/hr for the first 4 hours, then $310/hr or part thereof beyond
    that (overtime billed in whole-hour blocks).
    """
    hours = total_minutes / 60.0
    if hours <= NORMAL_RATE_HOURS:
        return hours * NORMAL_RATE_PER_HOUR
    overtime_hours = np.ceil(hours - NORMAL_RATE_HOURS)
    return NORMAL_RATE_HOURS * NORMAL_RATE_PER_HOUR + overtime_hours * OVERTIME_RATE_PER_HOUR


def two_opt(route, dur):
    """
    Simple 2-opt local search to reduce driving time for a fixed set of
    stores. Unload time doesn't depend on order, so only driving time
    needs improving here.

    2-opt is a local search: the local optimum it converges to depends on
    the order it's given as a starting point. Callers that are comparing
    several candidate orderings of the same store set should run this on
    every candidate BEFORE comparing/discarding any of them, not after
    picking a "best" one by raw (un-optimized) distance -- see
    `generate_route_pool.consider()`.
    """
    if len(route) < 3:
        return route

    best = list(route)
    best_time = route_driving_minutes(best, dur)
    improved = True
    while improved:
        improved = False
        for i in range(len(best) - 1):
            for j in range(i + 1, len(best)):
                candidate = best[:i] + best[i:j + 1][::-1] + best[j + 1:]
                cand_time = route_driving_minutes(candidate, dur)
                if cand_time < best_time - 1e-9:
                    best, best_time = candidate, cand_time
                    improved = True
    return best


def _construct_random_route(store_pool, demand, dur, rng, capacity, sanity_cutoff_min, top_k=3):
    """
    Build one route via randomized-greedy nearest-neighbour construction,
    returning every feasible partial route built along the way (not just
    the final one) so a single construction pass yields many pool
    candidates.
    """
    unvisited = list(store_pool)
    rng.shuffle(unvisited)
    start = unvisited[0]

    route = [start]
    pallets = demand[start]
    # driving time for Warehouse -> ... -> route[-1], NOT including the
    # return leg -- tracked incrementally so feasibility checks below are
    # O(1) instead of re-walking the whole route each time
    driving_out = dur.loc[WAREHOUSE, start] / 60.0
    snapshots = [list(route)]

    remaining = set(store_pool) - {start}

    while remaining:
        last = route[-1]
        feasible = []
        for s in remaining:
            new_pallets = pallets + demand[s]
            if new_pallets > capacity:
                continue
            trial_driving = driving_out + (dur.loc[last, s] + dur.loc[s, WAREHOUSE]) / 60.0
            trial_total = trial_driving + UNLOAD_MIN_PER_PALLET * new_pallets
            if trial_total > sanity_cutoff_min:
                continue
            feasible.append((dur.loc[last, s], s))

        if not feasible:
            break

        feasible.sort(key=lambda x: x[0])
        pick_from = feasible[:top_k]
        _, chosen = pick_from[rng.randrange(len(pick_from))]

        driving_out += dur.loc[last, chosen] / 60.0
        route.append(chosen)
        pallets += demand[chosen]
        remaining.discard(chosen)
        snapshots.append(list(route))

    return snapshots


def generate_route_pool(demand, dur, n_restarts=4000, capacity=CAPACITY_PALLETS,
                         sanity_cutoff_min=360, top_k=3, seed=0):
    """
    Build a large, deduplicated pool of candidate routes for one day-type's
    demand.

    Parameters
    ----------
    demand : dict[str, float]
        {store_name: pallets required}.
    dur : pd.DataFrame
        Warehouse<->store driving-time matrix (seconds), as returned by
        load_duration_matrix().
    n_restarts : int
        Number of randomized-greedy construction passes to run. Each pass
        contributes many candidates (every partial route along the way),
        so the final pool is much larger than n_restarts.
    capacity : float
        Truck capacity in pallets (hard constraint).
    sanity_cutoff_min : float
        Upper bound on total route time for a candidate to be kept, purely
        to bound pool size -- not a business rule. Singleton routes are
        exempt and always kept.
    top_k : int
        How many of the nearest feasible next-stores to randomize between
        at each construction step (higher = more random/diverse, lower =
        more purely greedy).
    seed : int
        RNG seed for reproducibility.

    Returns
    -------
    (pd.DataFrame, list[str])
        pool : one row per distinct set of stores (the best/shortest-driving
            ordering found for that set, after 2-opt). Columns: stores
            (tuple), n_stores, pallets, driving_min, unload_min, total_min,
            total_hours, cost, is_overtime.
        zero_demand_stores : stores with demand == 0 for this day-type,
            excluded from route construction entirely (see module
            docstring) and reported here rather than folded into the pool
            or flagged as "never covered".
    """
    all_stores = [s for s in demand if s != WAREHOUSE]
    stores = [s for s in all_stores if demand[s] > 0]
    zero_demand_stores = sorted(s for s in all_stores if demand[s] == 0)

    rng = random.Random(seed)

    best_for_set = {}  # frozenset(stores) -> (driving_min, ordered_route), both POST-2-opt

    def consider(route):
        # Optimize before comparing: 2-opt's result depends on the
        # starting order, so comparing/discarding on raw construction
        # time first would risk keeping an ordering that "looked" good
        # but optimizes worse than one that got thrown away.
        opt_route = two_opt(route, dur) if len(route) >= 3 else route
        key = frozenset(opt_route)
        driving_min = route_driving_minutes(opt_route, dur)
        if key not in best_for_set or driving_min < best_for_set[key][0] - 1e-9:
            best_for_set[key] = (driving_min, opt_route)

    # mandatory: every store with nonzero demand must have a singleton
    # candidate, regardless of the sanity cutoff, so the pool always
    # supports full coverage
    for s in stores:
        consider([s])

    for _ in range(n_restarts):
        for snap in _construct_random_route(stores, demand, dur, rng, capacity, sanity_cutoff_min, top_k):
            if len(snap) == 1:
                continue  # singletons already added above
            consider(snap)

    rows = []
    for key, (driving_min, route) in best_for_set.items():
        pallets = sum(demand[s] for s in route)
        unload_min = UNLOAD_MIN_PER_PALLET * pallets
        total_min = driving_min + unload_min
        rows.append({
            "stores": tuple(route),
            "n_stores": len(route),
            "pallets": pallets,
            "driving_min": round(driving_min, 1),
            "unload_min": unload_min,
            "total_min": round(total_min, 1),
            "total_hours": round(total_min / 60.0, 2),
            "cost": round(route_cost(total_min), 2),
            "is_overtime": total_min > NORMAL_RATE_HOURS * 60,
        })

    out = pd.DataFrame(rows).sort_values(["n_stores", "total_min"]).reset_index(drop=True)
    return out, zero_demand_stores


def pool_summary(pool_df, demand, zero_demand_stores=None):
    """Quick sanity-check summary of a generated route pool."""
    zero_demand_stores = zero_demand_stores or []
    n_stores_total = len([s for s in demand])
    covered = set()
    for stores in pool_df["stores"]:
        covered.update(stores)
    never_covered = set(demand) - covered - set(zero_demand_stores)
    return {
        "n_candidate_routes": len(pool_df),
        "n_stores_total": n_stores_total,
        "n_stores_covered_by_pool": len(covered),
        "stores_never_covered": sorted(never_covered),
        "zero_demand_stores_excluded": sorted(zero_demand_stores),
        "n_overtime_routes": int(pool_df["is_overtime"].sum()),
        "route_size_distribution": pool_df["n_stores"].value_counts().sort_index().to_dict(),
    }


def save_pools_to_excel(weekday_pool, saturday_pool, path="route_pools.xlsx"):
    """Write the Weekday and Saturday candidate route pools to a formatted .xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    body_font = Font(name="Arial")
    overtime_fill = PatternFill("solid", fgColor="FCE4E4")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, pool in [("Weekday Pool", weekday_pool), ("Saturday Pool", saturday_pool)]:
        ws = wb.create_sheet(sheet_name)
        display_cols = ["route", "n_stores", "pallets", "driving_min", "unload_min",
                         "total_min", "total_hours", "cost", "is_overtime"]
        display = pool.copy()
        display["route"] = display["stores"].apply(lambda s: "Warehouse -> " + " -> ".join(s) + " -> Warehouse")
        display = display[display_cols]

        for j, col in enumerate(display.columns, start=1):
            cell = ws.cell(row=1, column=j, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for i, row in enumerate(display.itertuples(index=False), start=2):
            is_overtime = row.is_overtime
            for j, val in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.font = body_font
                cell.border = border
                if is_overtime:
                    cell.fill = overtime_fill

        ws.column_dimensions["A"].width = 70
        for j in range(2, len(display_cols) + 1):
            ws.column_dimensions[get_column_letter(j)].width = 13
        ws.freeze_panes = "A2"

    wb.save(path)
    return path


from pathlib import Path

# store_routes.py is in Routes/, so the project root is one level up
PROJECT_ROOT = Path(__file__).resolve().parent.parent

DEMAND_XLSX = PROJECT_ROOT / "Demands" / "required_supply_weekday_saturday.xlsx"
DURATIONS_CSV = PROJECT_ROOT / "Resources" / "FoodstuffsDurations2026.csv"
ROUTE_POOLS_XLSX = PROJECT_ROOT / "Routes" / "route_pools.xlsx"


if __name__ == "__main__":
    weekday_demand, saturday_demand = load_demand(DEMAND_XLSX)
    dur = load_duration_matrix(DURATIONS_CSV)

    weekday_pool, weekday_zero = generate_route_pool(weekday_demand, dur, n_restarts=4000, seed=1)
    saturday_pool, saturday_zero = generate_route_pool(saturday_demand, dur, n_restarts=4000, seed=2)

    print("Weekday pool:", pool_summary(weekday_pool, weekday_demand, weekday_zero))
    print("Saturday pool:", pool_summary(saturday_pool, saturday_demand, saturday_zero))

    save_pools_to_excel(weekday_pool, saturday_pool, ROUTE_POOLS_XLSX)