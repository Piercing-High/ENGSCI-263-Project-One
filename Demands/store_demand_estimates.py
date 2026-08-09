"""
Foodstuffs demand analysis
===========================

Pipeline:
1. remove_outliers()      - flags outliers per store using a z-score computed
                             within each day-of-week group (so a Monday isn't
                             compared to a Sunday), and replaces them with the
                             median for that store/day-of-week.
2. remove_zero_columns()  - drops any date column that is zero for every
                             store (e.g. Sundays / public holidays when the
                             whole chain is closed).
3. supply_service_level() - given a per-store supply level, works out what
                             percentage of days that supply would have been
                             enough to cover demand, and (optionally) flags
                             whether that meets a target service level.

Assumptions (flag if these don't match what you had in mind):
- Data is wide format: one row per store, one column per date (dd/mm/yyyy).
- "Outlier" = a value whose z-score (computed against that store's other
  values on the same weekday) exceeds `z_thresh` (default 3.0).
- "Zero columns" = date columns that are 0 for every single store (not
  columns that are mostly zero).
- "Supply" is a per-store value you supply (e.g. units delivered/stocked).
  The function reports the fraction of (non-outlier, non-zero-day) days on
  which demand <= supply for that store.
"""

import pandas as pd
import numpy as np


def _weekday_groups(date_cols):
    """
    Group date-string columns by weekday (0=Monday ... 6=Sunday).

    Shared by every function that needs a Mon/Tue/.../Sat split, so the
    date parsing and grouping logic lives in exactly one place.
    """
    groups = {}
    for c in date_cols:
        wd = pd.to_datetime(c, dayfirst=True).weekday()
        groups.setdefault(wd, []).append(c)
    return groups


def _weekday_saturday_cols(date_cols):
    """
    Split date columns into a pooled Monday-Friday ("weekday") bucket and
    a Saturday bucket, ignoring Sunday. Shared by every function that
    needs the Weekday/Saturday split, so it's computed the same way (and
    only parsed) in exactly one place.
    """
    groups = _weekday_groups(date_cols)
    weekday_cols = [c for wd in range(5) for c in groups.get(wd, [])]
    saturday_cols = groups.get(5, [])
    return weekday_cols, saturday_cols


def remove_outliers(df, date_cols=None, z_thresh=3.5):
    """
    Replace outlier values with the median for the same store on the same
    day of week.

    Outliers are detected per store, within groups of columns that share a
    day of week (so e.g. all the Mondays for a given store are compared to
    each other, not to that store's Sundays which are naturally much lower).

    Uses the *modified* z-score (median + MAD, i.e. median absolute
    deviation) rather than the standard mean/std z-score. With only ~8
    values per store/weekday group, a single large outlier inflates the
    standard deviation enough to hide itself from a normal z-score test;
    MAD is far less sensitive to the outlier it's trying to detect.

    Parameters
    ----------
    df : pd.DataFrame
        Must have a 'Supermarket' column plus one column per date
        (parseable as dd/mm/yyyy).
    date_cols : list[str], optional
        Which columns to treat as demand/date columns. Defaults to all
        columns except 'Supermarket'.
    z_thresh : float
        Modified z-score magnitude above which a value is treated as an
        outlier. 3.5 is the commonly used default for this method.

    Returns
    -------
    pd.DataFrame
        Copy of df with outliers replaced.
    """
    df = df.copy()
    if date_cols is None:
        date_cols = [c for c in df.columns if c != "Supermarket"]
    df[date_cols] = df[date_cols].astype(float)

    # group columns by day-of-week
    weekday_groups = _weekday_groups(date_cols)

    for cols in weekday_groups.values():
        sub = df[cols]
        row_median = sub.median(axis=1)
        mad = sub.sub(row_median, axis=0).abs().median(axis=1).replace(0, np.nan)

        # modified z-score, per Iglewicz & Hoaglin
        modified_z = sub.sub(row_median, axis=0).div(mad, axis=0) * 0.6745

        is_outlier = modified_z.abs() > z_thresh

        # broadcast row_median across columns via axis=0 instead of
        # building a fully replicated same-shape replacement DataFrame
        df[cols] = sub.where(~is_outlier, row_median, axis=0)

    return df


def remove_zero_columns(df, date_cols=None):
    """
    Drop date columns that are zero for every store.

    Parameters
    ----------
    df : pd.DataFrame
    date_cols : list[str], optional
        Columns to check. Defaults to all columns except 'Supermarket'.

    Returns
    -------
    pd.DataFrame
        Copy of df with all-zero columns removed.
    """
    if date_cols is None:
        date_cols = [c for c in df.columns if c != "Supermarket"]

    zero_cols = [c for c in date_cols if (df[c] == 0).all()]
    return df.drop(columns=zero_cols)


def supply_service_level(df, supply, target_percent=None, date_cols=None):
    """
    For each store, work out what percentage of days a given supply level
    would have covered demand (demand <= supply).

    Parameters
    ----------
    df : pd.DataFrame
        Cleaned demand data (after remove_outliers + remove_zero_columns).
    supply : dict[str, float]
        Per-store supply level, keyed by the exact value in 'Supermarket'.
    target_percent : float, optional
        If given (e.g. 0.95 for 95%), an extra boolean column is added
        showing whether the store's achieved coverage meets this target.
    date_cols : list[str], optional
        Columns to treat as demand columns. Defaults to all columns except
        'Supermarket'.

    Returns
    -------
    pd.DataFrame
        Columns: Supermarket, supply, pct_days_demand_met
        (+ meets_target if target_percent is given)
    """
    if date_cols is None:
        date_cols = [c for c in df.columns if c != "Supermarket"]

    missing = set(df["Supermarket"]) - set(supply)
    if missing:
        raise KeyError(f"No supply value provided for store(s): {sorted(missing)}")

    supply_series = df["Supermarket"].map(supply)
    demand = df[date_cols].astype(float)
    pct_met = demand.le(supply_series, axis=0).mean(axis=1) * 100

    out = pd.DataFrame({
        "Supermarket": df["Supermarket"].values,
        "supply": supply_series.values,
        "pct_days_demand_met": pct_met.values,
    })
    if target_percent is not None:
        out["meets_target"] = out["pct_days_demand_met"] >= (target_percent * 100)
    return out


def required_supply_weekday_saturday(df, target_percent=0.95, date_cols=None):
    """
    For each store, work out the single supply amount that should be
    delivered on every weekday (Monday-Friday) and, separately, on every
    Saturday, on a standard delivery route -- such that demand is met at
    least `target_percent` of the time.

    This pools all Monday-Friday demand values for a store together and
    takes one quantile across the pooled set (i.e. it asks "if the same
    amount is delivered every weekday, what's the smallest amount that
    covers demand on target_percent of those weekday-days?"), and
    likewise pools all Saturdays into a second quantile. This matches a
    real delivery route, which drops the same load on a given weekday
    regardless of whether it happens to be a Monday or a Friday.

    Note: because this pools across weekdays rather than computing a
    separate figure per individual weekday-of-week, the target service
    level is met *on average across all weekdays combined* -- a
    particularly busy weekday (e.g. Friday) could still be under-supplied
    while quieter weekdays are over-supplied. If you need the target
    hit on every individual weekday, take the max of five per-weekday
    quantiles instead (the previous behaviour of this function).

    Sunday is assumed to already be dropped (all-zero / store closed) by
    remove_zero_columns() and is ignored here.

    Parameters
    ----------
    df : pd.DataFrame
        Cleaned demand data (after remove_outliers + remove_zero_columns).
    target_percent : float
        Desired service level, e.g. 0.95 for 95% of days covered.
    date_cols : list[str], optional
        Columns to treat as demand columns. Defaults to all columns except
        'Supermarket'.

    Returns
    -------
    pd.DataFrame
        Columns: Supermarket, Weekday, Saturday.
    """
    if date_cols is None:
        date_cols = [c for c in df.columns if c != "Supermarket"]

    weekday_cols, saturday_cols = _weekday_saturday_cols(date_cols)

    weekday_supply = (
        np.ceil(df[weekday_cols].astype(float).quantile(target_percent, axis=1))
        if weekday_cols else pd.Series(np.nan, index=df.index)
    )
    saturday_supply = (
        np.ceil(df[saturday_cols].astype(float).quantile(target_percent, axis=1))
        if saturday_cols else pd.Series(np.nan, index=df.index)
    )

    return pd.DataFrame({
        "Supermarket": df["Supermarket"].values,
        "Weekday": weekday_supply.values,
        "Saturday": saturday_supply.values,
    })


def collected_data_weekday_saturday(df, date_cols=None):
    """
    Group the (cleaned) raw demand values by store into just two buckets:
    "Weekday" (all Monday-Friday values pooled together) and "Saturday".

    Parameters
    ----------
    df : pd.DataFrame
        Cleaned demand data (after remove_outliers + remove_zero_columns).
    date_cols : list[str], optional
        Columns to treat as demand columns. Defaults to all columns except
        'Supermarket'.

    Returns
    -------
    dict
        {store_name: {"Weekday": [values, ...], "Saturday": [values, ...]}}
        Sunday is assumed already dropped and is ignored.
    """
    if date_cols is None:
        date_cols = [c for c in df.columns if c != "Supermarket"]

    weekday_cols, saturday_cols = _weekday_saturday_cols(date_cols)

    # select columns by name directly (pandas' column index is hash-based)
    # instead of manually resolving positions via date_cols.index(), which
    # would be an O(n*m) scan for n date columns and m selected columns
    weekday_values = df[weekday_cols].astype(float).values if weekday_cols else np.empty((len(df), 0))
    saturday_values = df[saturday_cols].astype(float).values if saturday_cols else np.empty((len(df), 0))
    stores = df["Supermarket"].values

    collected = {
        store: {
            "Weekday": weekday_values[i].tolist(),
            "Saturday": saturday_values[i].tolist(),
        }
        for i, store in enumerate(stores)
    }

    return collected


def print_collected_data_weekday_saturday(collected):
    """Print collected_data_weekday_saturday() output in 'Store: group: values' form."""
    for store, groups in collected.items():
        print(f"{store}:")
        for group, values in groups.items():
            print(f"  {group}: {values}")
        print()


def _write_collected_sheet(wb, collected):
    """Write the raw-values 'Collected Data' sheet (one block per store)."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    ws0 = wb.create_sheet("Collected Data", 0)
    store_font = Font(name="Arial", bold=True, size=12)
    day_font = Font(name="Arial", bold=True)
    value_font = Font(name="Arial")

    max_values_per_row = 1
    r = 1
    for store, groups in collected.items():
        cell = ws0.cell(row=r, column=1, value=store)
        cell.font = store_font
        r += 1
        for group, values in groups.items():
            ws0.cell(row=r, column=1, value=group).font = day_font
            for j, v in enumerate(values, start=2):
                c = ws0.cell(row=r, column=j, value=v)
                c.font = value_font
                c.number_format = "0"
            max_values_per_row = max(max_values_per_row, len(values))
            r += 1
        r += 1  # blank row between stores

    ws0.column_dimensions["A"].width = 14
    # size value columns off the actual widest row, not a hardcoded guess
    for j in range(2, max_values_per_row + 2):
        ws0.column_dimensions[get_column_letter(j)].width = 8


def _write_summary_sheet(ws, result_df):
    """Write the formatted 'Required Supply' heatmap table onto ws."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    n_rows, n_cols = result_df.shape

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    body_font = Font(name="Arial")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # header row
    for j, col in enumerate(result_df.columns, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # data rows
    for i, row in enumerate(result_df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = body_font
            cell.border = border
            if j > 1:
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")

    # column widths
    ws.column_dimensions["A"].width = 32
    for j in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 12

    ws.freeze_panes = "B2"

    # heatmap color scale over the weekday values
    last_col_letter = get_column_letter(n_cols)
    data_range = f"B2:{last_col_letter}{n_rows + 1}"
    ws.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type="min", start_color="FFF8696B",
            mid_type="percentile", mid_value=50, mid_color="FFFFEB84",
            end_type="max", end_color="FF63BE7B",
        ),
    )


def save_to_excel(result_df, path="required_supply_weekday_saturday.xlsx", collected=None):
    """
    Write a required_supply_weekday_saturday() result to a formatted
    .xlsx file: a bold header, autosized columns, and a heatmap-style
    color scale over the Weekday/Saturday values.

    If `collected` (output of collected_data_weekday_saturday()) is
    given, an extra "Collected Data" sheet is added first, listing the
    raw demand values behind each store's Weekday/Saturday summary
    figures — e.g.:

        Four Square Botany Junction
            Weekday     4  4  3  4  1  1  2  2  3  4
            Saturday    9  11  8

    Parameters
    ----------
    result_df : pd.DataFrame
        Output of required_supply_weekday_saturday().
    path : str
        Output file path (.xlsx).
    collected : dict, optional
        Output of collected_data_weekday_saturday().
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Required Supply"

    if collected is not None:
        _write_collected_sheet(wb, collected)

    _write_summary_sheet(ws, result_df)

    wb.save(path)
    return path


if __name__ == "__main__":
    # Example usage
    df_raw = pd.read_csv("Resources/FoodstuffsDemand2026.csv", encoding="utf-8-sig")

    df_clean = remove_outliers(df_raw, z_thresh=3.5)
    df_clean = remove_zero_columns(df_clean)

    result = required_supply_weekday_saturday(df_clean, target_percent=0.95)
    print(result.to_string(index=False))

    collected = collected_data_weekday_saturday(df_clean)
    print_collected_data_weekday_saturday(collected)

    save_to_excel(result, "required_supply_weekday_saturday.xlsx", collected=collected)