"""
Foodstuffs schedule simulation (Task 5)
=======================================

Monte Carlo simulation estimating the ACTUAL cost of running the
optimiser's fixed schedule under day-to-day demand variation and
Auckland traffic variation.

This one script does everything:
  1. Runs all four scenarios (weekday/Saturday x baseline/shedding).
  2. Runs a PAIRED comparison of weekday baseline vs shedding using common
     random numbers (same random day for both), with a paired t-test.
  3. Writes a formatted Excel workbook (summary, paired stats, raw costs).
  4. Saves a clean distribution plot (baseline vs shedding).

Method (per simulated day)
--------------------------
  1. Demand  - resample one observed value per store (bootstrap).
  2. Traffic - scale each route's driving time by a random multiplier;
               routes are pre-assigned to an 8am (peak) or 2pm (off-peak)
               shift, longest routes placed off-peak (greedy).
  3. Cost    - if a route's demand exceeds 16 pallets, run the whole route
               as a wet-lease (Linfox, $1400/2h block); else fleet model
               ($220/hr, $310/hr overtime beyond 4h).
  4. Sum     - total the day's cost; record wet-lease / overtime counts.

Assumptions / limitations (state in report)
--------------------------------------------
  - Demand resampled independently per store (bootstrap independence
    assumption; real demand may correlate across stores).
  - Single peak / off-peak multiplier range, NOT direction-specific. Real
    Auckland congestion is directional, so a blanket multiplier over/under-
    estimates individual legs, partly cancelling across a mixed route set.
  - Shift assignment fixed (greedy by route length); a full assignment
    optimisation is possible future work.

Run:  python "Simulation/simulation.py"
"""

import numpy as np
import pandas as pd
from pathlib import Path
from scipy import stats

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt


# ============================================================
# PATHS  (script in "Simulation/", repo root one level up)
# ============================================================

PROJECT_ROOT  = Path(__file__).resolve().parent.parent
DEMAND_CSV    = PROJECT_ROOT / "Resources" / "FoodstuffsDemand2026.csv"
DURATIONS_CSV = PROJECT_ROOT / "Resources" / "FoodstuffsDurations2026.csv"
SELECTED_XLSX = PROJECT_ROOT / "Linear Program" / "selected_routes.xlsx"
OUTPUT_XLSX   = PROJECT_ROOT / "Simulation" / "simulation_results.xlsx"
OUTPUT_PLOT   = PROJECT_ROOT / "Simulation" / "simulation_distribution.png"


# ============================================================
# CONSTANTS
# ============================================================

CAPACITY     = 16
UNLOAD_MIN   = 18
NORMAL_H     = 4.0
NORMAL_RATE  = 220.0
OT_RATE      = 310.0
LINFOX_BLOCK = 1400.0
WAREHOUSE    = "Warehouse"

PEAK_LO, PEAK_HI = 1.4, 1.8      # 8am peak multiplier range
OFF_LO,  OFF_HI  = 1.05, 1.15    # 2pm off-peak multiplier range

N_ITER = 5000     # 1000 to test, 5000 for final numbers
SEED   = 1

OPTIMISATION_ESTIMATE = 48327    # weekday baseline planning cost, for the plot

SCENARIOS = [
    ("Weekday Baseline",  "Weekday Baseline",  "weekday"),
    ("Saturday Baseline", "Saturday Baseline", "saturday"),
    ("Weekday Shedding",  "Weekday Shedding",  "weekday"),
    ("Saturday Shedding", "Saturday Shedding", "saturday"),
]


# ============================================================
# LOAD SHARED DATA
# ============================================================

dur = pd.read_csv(DURATIONS_CSV, index_col=0)

demand_df = pd.read_csv(DEMAND_CSV, encoding="utf-8-sig")
names = demand_df["Supermarket"].values
date_cols = [c for c in demand_df.columns if c != "Supermarket"]
demand_df[date_cols] = demand_df[date_cols].astype(float)

zero_cols = [c for c in date_cols if (demand_df[c] == 0).all()]
date_cols = [c for c in date_cols if c not in zero_cols]
weekday_cols  = [c for c in date_cols if pd.to_datetime(c, dayfirst=True).weekday() < 5]
saturday_cols = [c for c in date_cols if pd.to_datetime(c, dayfirst=True).weekday() == 5]

HIST = {
    "weekday":  {names[i]: demand_df.loc[i, weekday_cols].values.astype(int)
                 for i in range(len(names))},
    "saturday": {names[i]: demand_df.loc[i, saturday_cols].values.astype(int)
                 for i in range(len(names))},
}


# ============================================================
# HELPERS
# ============================================================

def driving_minutes(stops):
    if not stops:
        return 0.0
    seq = [WAREHOUSE] + stops + [WAREHOUSE]
    return sum(dur.loc[a, b] for a, b in zip(seq, seq[1:])) / 60.0


def fleet_cost(total_min):
    hours = total_min / 60.0
    if hours <= NORMAL_H:
        return hours * NORMAL_RATE
    return NORMAL_H * NORMAL_RATE + np.ceil(hours - NORMAL_H) * OT_RATE


def linfox_cost(total_min):
    return np.ceil(total_min / 60.0 / 2.0) * LINFOX_BLOCK


def prep_routes(sheet):
    """Load a scenario's routes, compute base drive time, assign AM/PM."""
    r = pd.read_excel(SELECTED_XLSX, sheet_name=sheet)
    r["stops"] = r["route"].apply(lambda x: x.split(" -> ")[1:-1])
    r["base_drive"] = r["stops"].apply(driving_minutes)
    r = r.sort_values("base_drive", ascending=False).reset_index(drop=True)
    half = len(r) // 2
    r["shift"] = ["PM"] * half + ["AM"] * (len(r) - half)   # longest -> PM
    return r


# ============================================================
# SIMULATION
# ============================================================

def route_set_cost(routes, demand, traffic):
    """Cost of one route set given a FIXED demand dict and per-shift traffic."""
    total, n_wet, n_ot = 0.0, 0, 0
    for _, r in routes.iterrows():
        pallets = sum(demand[s] for s in r["stops"])
        total_min = r["base_drive"] * traffic[r["shift"]] + UNLOAD_MIN * pallets
        if pallets > CAPACITY:
            total += linfox_cost(total_min); n_wet += 1
        else:
            total += fleet_cost(total_min)
            if total_min > NORMAL_H * 60:
                n_ot += 1
    return total, n_wet, n_ot


def run_scenario(routes, hist, n_iter, seed):
    """Independent Monte Carlo run for one scenario."""
    rng = np.random.default_rng(seed)
    costs = np.empty(n_iter); wets = np.empty(n_iter); ots = np.empty(n_iter)
    for i in range(n_iter):
        demand = {s: int(rng.choice(v)) for s, v in hist.items()}
        traffic = {"AM": rng.uniform(PEAK_LO, PEAK_HI),
                   "PM": rng.uniform(OFF_LO, OFF_HI)}
        costs[i], wets[i], ots[i] = route_set_cost(routes, demand, traffic)
    return costs, wets, ots


def run_paired(routes_a, routes_b, hist, n_iter, seed):
    """
    Paired comparison using common random numbers: the SAME random demand
    and traffic are used for both route sets each day, so the comparison
    isolates the policy difference rather than day-to-day noise.
    """
    rng = np.random.default_rng(seed)
    a = np.empty(n_iter); b = np.empty(n_iter)
    for i in range(n_iter):
        demand = {s: int(rng.choice(v)) for s, v in hist.items()}
        traffic = {"AM": rng.uniform(PEAK_LO, PEAK_HI),
                   "PM": rng.uniform(OFF_LO, OFF_HI)}
        a[i] = route_set_cost(routes_a, demand, traffic)[0]
        b[i] = route_set_cost(routes_b, demand, traffic)[0]
    return a, b


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":

    # ---- 1. all four scenarios (independent runs) ----
    summary_rows = []
    raw_costs = {}
    for label, sheet, day_type in SCENARIOS:
        routes = prep_routes(sheet)
        costs, wets, ots = run_scenario(routes, HIST[day_type], N_ITER, seed=SEED)
        raw_costs[label] = costs
        summary_rows.append({
            "Scenario":          label,
            "Iterations":        N_ITER,
            "Mean cost ($)":     round(costs.mean()),
            "Std dev ($)":       round(costs.std()),
            "CI low 2.5% ($)":   round(np.percentile(costs, 2.5)),
            "CI high 97.5% ($)": round(np.percentile(costs, 97.5)),
            "Min ($)":           round(costs.min()),
            "Max ($)":           round(costs.max()),
            "Wet-lease days %":  round((wets > 0).mean() * 100, 1),
            "Avg wet-lease/day": round(wets.mean(), 3),
            "Avg overtime/day":  round(ots.mean(), 1),
        })
        print(f"Done: {label:20s} mean ${costs.mean():,.0f}")
    summary = pd.DataFrame(summary_rows)

    # ---- 2. paired comparison: weekday baseline vs shedding ----
    base_r = prep_routes("Weekday Baseline")
    shed_r = prep_routes("Weekday Shedding")
    base_p, shed_p = run_paired(base_r, shed_r, HIST["weekday"], N_ITER, seed=SEED)
    diff = base_p - shed_p       # saving from shedding, per identical day
    t_stat, p_val = stats.ttest_rel(base_p, shed_p)
    paired = pd.DataFrame([{
        "Comparison":            "Weekday: Baseline vs Shedding (common random numbers)",
        "Baseline mean ($)":     round(base_p.mean()),
        "Shedding mean ($)":     round(shed_p.mean()),
        "Mean saving/day ($)":   round(diff.mean()),
        "Saving CI low ($)":     round(np.percentile(diff, 2.5)),
        "Saving CI high ($)":    round(np.percentile(diff, 97.5)),
        "Days shedding cheaper %": round((diff > 0).mean() * 100, 1),
        "Paired t-statistic":    round(t_stat, 1),
        "p-value":               f"{p_val:.2e}",
    }])
    print(f"\nPaired mean saving from shedding: ${diff.mean():,.0f}/day "
          f"(95% CI ${np.percentile(diff,2.5):,.0f}-${np.percentile(diff,97.5):,.0f})")

    # ---- 3. write formatted Excel ----
    raw = pd.DataFrame(raw_costs)
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        paired.to_excel(writer, sheet_name="Paired Comparison", index=False)
        raw.to_excel(writer, sheet_name="Raw daily costs", index=False)

        # light formatting: bold headers, autofit-ish column widths, freeze header
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F3A5F")
        for sheet_name in ["Summary", "Paired Comparison", "Raw daily costs"]:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
            for col in ws.columns:
                width = max((len(str(c.value)) for c in col if c.value is not None),
                            default=10)
                ws.column_dimensions[col[0].column_letter].width = min(width + 3, 40)

    print(f"Saved results to: {OUTPUT_XLSX}")

    # ---- 4. distribution plot (clean, nothing overlapping) ----
    fig, ax = plt.subplots(figsize=(11, 6), dpi=200)
    fig.patch.set_facecolor("white"); ax.set_facecolor("white")

    base_costs = raw_costs["Weekday Baseline"]
    shed_costs = raw_costs["Weekday Shedding"]
    ax.hist(base_costs, bins=45, color="#2a78d6", alpha=0.8,
            label="Baseline (all stores served)", edgecolor="white", lw=0.3)
    ax.hist(shed_costs, bins=45, color="#eb6834", alpha=0.8,
            label="Fuel-reduction (20% shedding)", edgecolor="white", lw=0.3)

    ymax = ax.get_ylim()[1]
    ax.set_ylim(0, ymax * 1.15)   # headroom for labels above bars
    for data, col in [(shed_costs, "#993C1D"), (base_costs, "#185FA5")]:
        mu = data.mean()
        ax.axvline(mu, color=col, lw=2)
        ax.annotate(f"${mu:,.0f}", xy=(mu, ymax * 1.02), xytext=(-6, 0),
                    textcoords="offset points", ha="right",
                    fontsize=11, fontweight="bold", color=col)

    ax.axvline(OPTIMISATION_ESTIMATE, color="#555", lw=1.6, ls="--")
    ax.annotate(f"Optimisation\nestimate ${OPTIMISATION_ESTIMATE:,}",
                xy=(OPTIMISATION_ESTIMATE, ymax * 0.55), xytext=(-12, 0),
                textcoords="offset points", ha="right", fontsize=9.5,
                color="#555", linespacing=1.3)

    ax.set_xlabel("Simulated daily cost", fontsize=12)
    ax.set_ylabel(f"Number of days (out of {N_ITER})", fontsize=12)
    ax.set_title("Distribution of actual weekday cost under demand & traffic variation",
                 fontsize=14, fontweight="bold", color="#1F3A5F", pad=14, loc="left")
    ax.legend(fontsize=10.5, loc="upper right", bbox_to_anchor=(0.995, 0.99),
              framealpha=0.95, edgecolor="#ddd")
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(axis="y", color="#e8e6df", lw=0.6); ax.set_axisbelow(True)
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"${x/1000:.0f}k"))
    plt.tight_layout()
    fig.savefig(OUTPUT_PLOT, dpi=200, facecolor="white", bbox_inches="tight")
    print(f"Saved plot to:    {OUTPUT_PLOT}")

    print("\n" + "=" * 62)
    print(summary.to_string(index=False))
    print("=" * 62)
